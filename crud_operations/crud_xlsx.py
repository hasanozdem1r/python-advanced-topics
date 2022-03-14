# PACKAGE IMPORTS
from pathlib import Path
import openpyxl as oxl

file = Path("data_sources", "VOCABULARY.xlsx")


class XlsxOperations:
    def __init__(self, file_path: str) -> None:
        self._file_path = file_path
        self._workbook_obj = oxl.load_workbook(self._file_path)
        self._sheet_obj = self._workbook_obj.active

    def read_all_xlsx(self) -> list:
        """
        Read all data from sheet
        :return result_df:<list:str> after read data to process return as list
        """
        max_row = self._sheet_obj.max_row
        result_df: list = []
        for item in range(1, max_row + 1):
            cell_obj = self._sheet_obj.cell(row=item, column=1)
            result_df.append(str(cell_obj.value))
        # remove duplicates
        result_df = list(set(result_df))
        return result_df

    def write_all_xlsx(self, data: list, new_sheet_name: str) -> None:
        """
        This method create and excel sheet with given sheet name and write given data to a new sheet
        :param data:<list:str> list of data
        :param new_sheet_name:<str> sheet name for creating new
        :return:
        """
        workbook = oxl.Workbook()
        worksheet = workbook.active
        for row, item in enumerate(data):
            col_cell = "A"
            worksheet[col_cell + str(row + 1)] = str(item)
        workbook.save(f"{new_sheet_name}.xlsx")


xlsx_obj = XlsxOperations(file)
result_data: list = xlsx_obj.read_all_xlsx()
xlsx_obj.write_all_xlsx(result_data, "vocabulary")
